import string
import spacy
from nltk.corpus import stopwords
import os
from openpyxl import load_workbook
from matplotlib import pyplot as plt
from zemberek import TurkishMorphology
from collections import defaultdict
import itertools
import networkx as nx

import tkinter as tk
from tkinter import filedialog
import sys



def turkish_to_english(text):
    turkish_chars = {
        'ş': 's', 'ç': 'c', 'ğ': 'g', 'ü': 'u', 'ö': 'o', 'ı': 'i', 'İ': 'I',
        'Ş': 'S', 'Ç': 'C', 'Ğ': 'G', 'Ü': 'U', 'Ö': 'O'
    }

    for turkish_char, english_char in turkish_chars.items():
        text = text.replace(turkish_char, english_char)

    return text.upper()


def extract_special_words(sentence):
    # Tokenize and analyze the sentence
    analysis_results = morphology.analyze_sentence(sentence)
    disambiguation_results = morphology.disambiguate(sentence, analysis_results)

    for word_analysis in disambiguation_results.best_analysis():
        lemma = word_analysis.item.normalized_lemma()  # Correctly call the method
        primary_pos = str(word_analysis.item.primary_pos)
        secondary_pos = str(word_analysis.item.secondary_pos)

        if primary_pos == "PrimaryPos.Noun" and lemma[0].isupper():
            return lemma


def remove_stopwords(text_list, stopwords):
    final_sentence_list = []

    for sentence in text_list:
        splitted_sentence = sentence.split(" ")
        for stopword in stopwords:
            a = 0
            for word in splitted_sentence:
                if word == stopword:
                    splitted_sentence[a] = ''
                a += 1

        cleaned_sentence = [word for word in splitted_sentence if word != '']
        final_sentence_list.append(' '.join(cleaned_sentence))

    return final_sentence_list


def remove_punctuation(text):
    # Yaygın özel karakterlerin listesini oluşturuyoruz
    additional_characters = '“”‘’–—•…«»©®'
    # string.punctuation içindeki karakterlere ekliyoruz
    punctuation = string.punctuation + additional_characters
    # Her bir işaret için boşluk ile değiştirme işlemi yapıyoruz
    for char in punctuation:
        text = text.replace(char, " ")
    return text


def removePunctuationsInSentences(list):
    a = 0
    for sentence in list:
        list[a] = remove_punctuation(sentence)
        a += 1
    return list


def read_first_words(file_name):
    with open(file_name, 'r') as file:
        word_list = [line.split(' ')[0] for line in file]
    return word_list


def build_character_graph(characters, sentences):
    # Sadece 3 ve daha fazla harfli karakter isimlerini filtrele
    filtered_characters = [char for char in characters if len(char) >= 3]

    # Boş bir yönsüz grafik oluştur
    G = nx.Graph()

    # Filtrelenmiş karakterleri düğüm olarak ekle
    for character in filtered_characters:
        G.add_node(character)

    # Her cümledeki karakterler arasındaki ilişkileri işleme al
    for sentence in sentences:
        # Bu cümlede geçen filtrelenmiş karakterleri bul
        found_characters = [char for char in filtered_characters if char in sentence]

        # Eşzamanlı geçen her karakter çifti için grafiğe kenar ekle veya ağırlığını artır
        for pair in itertools.combinations(found_characters, 2):
            if G.has_edge(*pair):
                # Eğer kenar zaten varsa ağırlığını artır
                G[pair[0]][pair[1]]['weight'] += 1
            else:
                # Kenar yoksa, yeni bir kenar ekle ve ağırlığını 1 yap
                G.add_edge(pair[0], pair[1], weight=1)

    return G


def sort_edges_by_weight(G):
    # Kenarları ağırlıklarına göre büyükten küçüğe sıralama
    sorted_edges = sorted(G.edges(data=True), key=lambda x: x[2]['weight'], reverse=True)

    # Sıralanmış kenarları içeren yeni bir graf oluşturma
    sorted_G = nx.Graph()
    sorted_G.add_nodes_from(G.nodes())
    sorted_G.add_edges_from((edge[0], edge[1], {'weight': edge[2]['weight']}) for edge in sorted_edges)

    return sorted_G


def file_creater(graph, fileName):
    with open(fileName, "w") as file:
        for edge in graph.edges(data=True):
            file.write(f"{edge[0]} - {edge[1]}  {edge[2]['weight']} \n")
        file.close()


def compute_centrality_measures(graph):
    centrality_measures = {}

    # Degree Centrality
    centrality_measures['degree_centrality'] = nx.degree_centrality(graph)

    # Betweenness Centrality
    centrality_measures['betweenness_centrality'] = nx.betweenness_centrality(graph)

    # Closeness Centrality
    centrality_measures['closeness_centrality'] = nx.closeness_centrality(graph)

    # Eigenvector Centrality
    centrality_measures['eigenvector_centrality'] = nx.eigenvector_centrality(graph)

    return centrality_measures


def print_top_10_centralities(centrality_measures):
    for measure_name, values in centrality_measures.items():
        # Değerleri büyükten küçüğe sıralama
        sorted_values = sorted(values.items(), key=lambda item: item[1], reverse=True)
        print(f"\nTop 10 {measure_name.capitalize()} Nodes:")
        for node, value in sorted_values[:10]:
            print(f"Node: {node}, Centrality: {value:.4f}")


def draw_graphs_side_by_side(graph1, graph2, title1, title2):
    # Yan yana iki grafik için subplot oluşturma
    fig, axes = plt.subplots(1, 2, figsize=(18, 8))  # 1 satır, 2 sütun

    # Birinci grafiği çizme
    pos1 = nx.spring_layout(graph1, seed=42)
    nx.draw(graph1, pos1, ax=axes[0], with_labels=True, node_color='skyblue', edge_color='gray', node_size=3000,
            font_size=12, font_weight='bold')
    axes[0].set_title(title1)

    # İkinci grafiği çizme
    pos2 = nx.spring_layout(graph2, seed=42)
    nx.draw(graph2, pos2, ax=axes[1], with_labels=True, node_color='lightgreen', edge_color='gray', node_size=3000,
            font_size=12, font_weight='bold')
    axes[1].set_title(title2)

    # Grafikler arası boşluk ve başlık ayarı
    plt.tight_layout()
    if os.environ.get('DISPLAY', '') == '':
        print("No display found. Saving figure instead of showing.")
        plt.savefig('graph_output.png')
    else:
        plt.show()


def save_centrality_measures_to_file(graphs, filename):
    with open(filename, 'w') as file:
        for name, graph in graphs.items():
            file.write(f"\nCentrality Measures for {name}:\n")
            centrality_measures = compute_centrality_measures(graph)
            for measure_name, values in centrality_measures.items():
                sorted_values = sorted(values.items(), key=lambda item: item[1], reverse=True)
                file.write(f"\nTop 10 {measure_name.capitalize()} Nodes:\n")
                for node, value in sorted_values[:10]:
                    file.write(f"Node: {node}, Centrality: {value:.4f}\n")


def main():
    flag = False
    root = tk.Tk()
    root.withdraw()  # Ana pencereyi gizle

    # Select the Excel file first
    print("Please select the Excel file containing the book (Col 1: TR, Col 2: EN).")
    input_file_path = filedialog.askopenfilename(title="Select Book Excel File", filetypes=(("Excel files", "*.xlsx"),))
    
    if not input_file_path:
        print("No file selected. Exiting.")
        return

    input_filename = os.path.basename(input_file_path)
    book_name = os.path.splitext(input_filename)[0]
    output_dir = os.path.join(os.getcwd(), f"Results_{book_name}")
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created output directory: {output_dir}")
    else:
        print(f"Using existing output directory: {output_dir}")

    charactersTRListFromUser = list()
    charactersENListFromUser = list()
    #Taking input correctly
    while True:
        try:
            choice = input(
                'If you already have the character files (charactersTR.txt and charactersEN.txt), type "1". If not, type "2" to extract and process the characters.\n')
        except EOFError:
            print("No input provided (EOF). Defaulting to option 2 (extract and process).")
            choice = '2'

        #User has character txt
        if choice == '1':
            tr_file_path = filedialog.askopenfilename(title="Select charactersTR.txt", filetypes=(("Text files", "*.txt"),))
            en_file_path = filedialog.askopenfilename(title="Select charactersEN.txt", filetypes=(("Text files", "*.txt"),))

            if tr_file_path and en_file_path:
                charactersTRListFromUser = read_first_words(tr_file_path)
                charactersENListFromUser = read_first_words(en_file_path)
                break  # break the loop when files are correctly selected
        # User has no character txt
        elif choice == '2':
            flag = True
            break
        else:
            print('Please enter only "1" or "2".')

    # Initialize Zemberek Turkish Morphology
    morphology = TurkishMorphology.create_with_defaults()

    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        namesToCheck = open(os.path.join(script_dir, 'turkish_names.txt'), 'r')
    except FileNotFoundError:
        print("turkish_names.txt not found in script directory. Please ensure it exists.")
        return

    nameList = list()
    # extracting turkish special names
    for line in namesToCheck:
        tempName = line.split(' ')
        nameList.append(tempName[0])
    namesToCheck.close()

    # opening book on xlsx
    try:
        wb = load_workbook(input_file_path)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    sheets = wb.sheetnames
    textTR = list()
    textEN = list()
    # Reading book
    for sheet_name in sheets:
        sheet = wb[sheet_name]
        rows = sheet.max_row
        cols = sheet.max_column

        for i in range(3, rows + 1):
            for j in range(1, cols):

                cell_value = sheet.cell(row=i, column=j).value
                if j == 1 and cell_value:
                    textTR.append(cell_value)
                elif j == 2 and cell_value:
                    textEN.append(cell_value)

    stopwordsEN = stopwords.words('english')  # taking english stopwords
    stopwordsTR = stopwords.words('turkish')  # taking turkish stopwords
    stopwordsEN.append('The')
    stopwordsTR.append('Bu')
    stopwordsTR.append('O')

    final_sentenceTRList = remove_stopwords(textTR, stopwordsTR)  # removing turkish stopwords from sentences
    final_sentenceENList = remove_stopwords(textEN, stopwordsEN)  # removing english stopwords from sentences

    final_sentenceTRList = removePunctuationsInSentences(final_sentenceTRList)  # removing punctuations from turkish sentences
    final_sentenceENList = removePunctuationsInSentences(final_sentenceENList)  # removing punctuations from english sentences
    # removing stopwords again to avoid mistaken lemmatization
    sentenceTRListFinal = remove_stopwords(final_sentenceTRList, stopwordsTR)
    sentenceENListFinal = remove_stopwords(final_sentenceENList, stopwordsEN)

    nlp = spacy.load("en_core_web_sm")
    lemmasEN = list()
    lemmasTR = list()
    lemmaTRGraphList = list()
    tempTRList = list()
    # Lemmatization in English
    for lem in sentenceENListFinal:
        doc = nlp(lem)
        lemmas = [token.lemma_ for token in doc]
        lemmasEN.append(' '.join(lemmas))
    # Lemmatization in Turkish
    for sentence in sentenceTRListFinal:
        analysis = morphology.analyze_sentence(sentence)
        after = morphology.disambiguate(sentence, analysis)
        for s in after.best_analysis():
            stem = s.get_stem()
            lemmasTR.append(stem)
            tempTRList.append(stem)

        lemmaTRGraphList.append(' '.join(tempTRList))
        tempTRList.clear()
    itemCount = defaultdict(int)

    # extracting and counting turkish characters
    for sentence in sentenceTRListFinal:
        sentenceList = sentence.split(' ')
        for word in sentenceList:
            if word[0].isupper():
                for item in nameList:
                    if item == turkish_to_english(word):
                        itemCount[word] += 1
    sortedItem = sorted(itemCount.items(), key=lambda x: x[1], reverse=True)

    # writing turkish characters into txt file
    with open(os.path.join(output_dir, "charactersTR.txt"), "w") as file:
        for label, count in sortedItem:
            file.write(f"{label}  {count} \n")
        file.close()

    # extracting and counting english characters
    undesired_labels = {"DATE", "ORDINAL", "CARDINAL", "TIME", "GPE"} # avoiding undesired type of characters
    label_counts = defaultdict(int)
    for sentence in lemmasEN:
        doc = nlp(sentence)
        for ent in doc.ents:
            if ent.label_ not in undesired_labels:
                label_counts[ent.text] += 1
    sorted_label_counts = sorted(label_counts.items(), key=lambda x: x[1], reverse=True)

    # writing english characters into txt file
    with open(os.path.join(output_dir, "charactersEN.txt"), "w") as file:
        for label, count in sorted_label_counts:
            file.write(f"{label}  {count} \n")
        file.close()
    countWordTr = defaultdict(int)

    # extracting and counting turkish words
    for words in lemmasTR:
        countWordTr[words] += 1
    sortedWordTr = sorted(countWordTr.items(), key=lambda x: x[1], reverse=True)
    # writing turkish words into txt file
    with open(os.path.join(output_dir, "wordsTR.txt"), "w") as file:
        for label, count in sortedWordTr:
            file.write(f"{label}  {count} \n")
        file.close()

    # extracting and counting english words
    countWordEN = defaultdict(int)
    for sentence in lemmasEN:
        splittedSentence = sentence.split(' ')
        for words in splittedSentence:
            countWordEN[words] += 1
    sortedWordEN = sorted(countWordEN.items(), key=lambda x: x[1], reverse=True)
    # writing english words into txt file
    with open(os.path.join(output_dir, "wordsEN.txt"), "w") as file:
        for label, count in sortedWordEN:
            file.write(f"{label}  {count} \n")
        file.close()

    # state that user doesn't have the character txt
    if flag:
        # reading characters from newly created files in output_dir
        charactersTRList = read_first_words(os.path.join(output_dir, 'charactersTR.txt')) 
        charactersENList = read_first_words(os.path.join(output_dir, 'charactersEN.txt')) 

        # reading words from newly created files in output_dir
        wordsTRList = read_first_words(os.path.join(output_dir, 'wordsTR.txt'))
        wordsENList = read_first_words(os.path.join(output_dir, 'wordsEN.txt'))

        # building the relation graph and sorting by their weights
        charactersGraphTR = sort_edges_by_weight(build_character_graph(charactersTRList, sentenceTRListFinal))
        charactersGraphEN = sort_edges_by_weight(build_character_graph(charactersENList, lemmasEN))
        wordsGraphTR = sort_edges_by_weight(build_character_graph(wordsTRList, lemmaTRGraphList))
        wordsGraphEN = sort_edges_by_weight(build_character_graph(wordsENList, lemmasEN))

        # writing relation graphs into txt
        file_creater(wordsGraphTR, os.path.join(output_dir, "wordsRelationsTR.txt"))
        file_creater(wordsGraphEN, os.path.join(output_dir, "wordsRelationsEN.txt"))
        file_creater(charactersGraphTR, os.path.join(output_dir, "characterRelationsTR.txt"))
        file_creater(charactersGraphEN, os.path.join(output_dir, "characterRelationsEN.txt"))

        twoGraphs = {
            "charactersGraphTR": charactersGraphTR,
            "charactersGraphEN": charactersGraphEN,
        }
        graphs = {
            "charactersGraphTR": charactersGraphTR,
            "charactersGraphEN": charactersGraphEN,
            "wordsGraphTR": wordsGraphTR,
            "wordsGraphEN": wordsGraphEN
        }

        # writing centrality measures into txt
        save_centrality_measures_to_file(graphs, os.path.join(output_dir, "centrality_measures.txt"))

        # plotting two graphs side by side
        draw_graphs_side_by_side(twoGraphs["charactersGraphTR"], twoGraphs["charactersGraphEN"], "Characters Graph TR",
                                 "Characters Graph EN")
    # state that user already has the character txt
    else:
        # reading words from txt - Assuming they are in output_dir if flow is correct, but here we can't assume.
        # However, option 2 generates them. Option 1 asks for chars. 
        # But 'wordsTR.txt' etc are generated in this run regardless of option 1?
        # Looking at code: yes, "writing turkish words into txt file" happens before the if flag block.
        # So they are in output_dir.
        
        wordsTRList = read_first_words(os.path.join(output_dir, 'wordsTR.txt'))
        wordsENList = read_first_words(os.path.join(output_dir, 'wordsEN.txt'))

        # building the relation graph and sorting by their weights
        charactersTRListFromUser = sort_edges_by_weight(build_character_graph(charactersTRListFromUser, sentenceTRListFinal))
        charactersENListFromUser = sort_edges_by_weight(build_character_graph(charactersENListFromUser, lemmasEN))
        wordsGraphTR = sort_edges_by_weight(build_character_graph(wordsTRList, lemmaTRGraphList))
        wordsGraphEN = sort_edges_by_weight(build_character_graph(wordsENList, lemmasEN))

        # writing relation graphs into txt
        file_creater(wordsGraphTR, os.path.join(output_dir, "wordsRelationsTR.txt"))
        file_creater(wordsGraphEN, os.path.join(output_dir, "wordsRelationsEN.txt"))
        file_creater(charactersTRListFromUser, os.path.join(output_dir, "characterRelationsTRFromUser.txt"))
        file_creater(charactersENListFromUser, os.path.join(output_dir, "characterRelationsENFromUser.txt"))
        twoGraphs = {
            "charactersGraphTRFromUser": charactersTRListFromUser,
            "charactersGraphENFromUser": charactersENListFromUser,
        }
        graphs = {
            "charactersGraphTRFromUser": charactersTRListFromUser,
            "charactersGraphENFromUser": charactersENListFromUser,
            "wordsGraphTR": wordsGraphTR,
            "wordsGraphEN": wordsGraphEN
        }

        # writing centrality measures into txt
        save_centrality_measures_to_file(graphs, os.path.join(output_dir, "centrality_measuresFromUser.txt"))

        # plotting two graphs side by side
        draw_graphs_side_by_side(twoGraphs["charactersGraphTRFromUser"], twoGraphs["charactersGraphENFromUser"],
                                 "Characters Graph TR",
                                 "Characters Graph EN")

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == '--no-gui':
       # Mock tk or handle no gui if needed, but for now just run main
       pass
    main()
